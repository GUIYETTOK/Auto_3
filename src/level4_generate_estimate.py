from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
import xlrd
from xlutils.copy import copy as xl_copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from zipfile import BadZipFile

from excel_utils import REQUIRED_HEADERS, find_header_row, get_file_datetime, normalize_text
from level1_build_db import iter_estimate_files
from level2_parse_request import iter_request_files, parse_request
from level3_match_prices import match_prices


def pick_latest_template(db_folder: Path) -> Path:
    candidates = list(iter_estimate_files(db_folder))
    if not candidates:
        raise SystemExit("견적서 템플릿을 찾지 못했습니다.")
    xlsx_candidates = [c for c in candidates if c.suffix.lower() == ".xlsx"]
    if xlsx_candidates:
        xlsx_candidates.sort(key=get_file_datetime, reverse=True)
        return xlsx_candidates[0]
    candidates.sort(key=get_file_datetime, reverse=True)
    return candidates[0]


def find_extra_columns(df: pd.DataFrame, header_row: int) -> Dict[str, int]:
    row = df.iloc[header_row]
    normalized = [normalize_text(v) for v in row.tolist()]
    col_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(normalized):
        if cell == "번호":
            col_map["번호"] = col_idx
    return col_map


def compute_spans(
    df: pd.DataFrame, header_row: int
) -> Dict[str, Tuple[int, int]]:
    row = df.iloc[header_row]
    normalized = [normalize_text(v) for v in row.tolist()]
    labels = {
        "번호",
        "품명",
        "규격",
        "단위",
        "수량",
        "단가",
        "금액",
        "비고",
    }
    markers = [(idx, cell) for idx, cell in enumerate(normalized) if cell in labels]
    if not markers:
        return {}
    markers.sort(key=lambda x: x[0])
    max_col = df.shape[1] - 1
    spans: Dict[str, Tuple[int, int]] = {}
    for i, (start_idx, label) in enumerate(markers):
        end_idx = max_col
        if i + 1 < len(markers):
            end_idx = markers[i + 1][0] - 1
        spans[label] = (start_idx, end_idx)
    return spans


def find_header_row_openpyxl(
    ws,
) -> Optional[Tuple[int, Dict[str, int]]]:
    for row in ws.iter_rows():
        normalized = [normalize_text(cell.value) for cell in row]
        header_map: Dict[str, int] = {}
        for idx, cell in enumerate(normalized, start=1):
            if cell in REQUIRED_HEADERS and cell not in header_map:
                header_map[cell] = idx
            if cell == "번호" and "번호" not in header_map:
                header_map["번호"] = idx
        if all(h in header_map for h in REQUIRED_HEADERS):
            return row[0].row, header_map
    return None


def compute_request_label(input_path: Path) -> str:
    if input_path.is_file():
        stem = input_path.stem
    else:
        files = list(iter_request_files(input_path))
        if not files:
            return ""
        files.sort(key=get_file_datetime, reverse=True)
        stem = files[0].stem
    compact = "".join(stem.split())
    return compact[:4]


def find_last_data_row(ws, start_row: int, start_col: int, end_col: int) -> int:
    empty_streak = 0
    last_row = start_row
    for row_idx in range(start_row, ws.max_row + 1):
        has_any = False
        for col_idx in range(start_col, end_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                has_any = True
                break
        if has_any:
            empty_streak = 0
            last_row = row_idx
        else:
            empty_streak += 1
            if empty_streak >= 2:
                break
    return max(last_row, start_row)


def resolve_writable_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return cell


def clear_existing_rows(
    sheet,
    header_row: int,
    spans: Dict[str, Tuple[int, int]],
    max_row: int,
) -> None:
    for row_idx in range(header_row + 1, max_row + 1):
        for key in ["번호", "품명", "규격", "단위", "수량", "단가", "금액"]:
            if key not in spans:
                continue
            start_col, end_col = spans[key]
            for col in range(start_col, end_col + 1):
                sheet.write(row_idx, col, "")


def write_rows(
    sheet,
    start_row: int,
    spans: Dict[str, Tuple[int, int]],
    matched_rows,
) -> None:
    for idx, matched in enumerate(matched_rows, start=1):
        row_idx = start_row + (idx - 1)
        if "번호" in spans:
            sheet.write(row_idx, spans["번호"][0], idx)
        if "품명" in spans:
            sheet.write(row_idx, spans["품명"][0], matched.request.품명)
        if "규격" in spans:
            sheet.write(row_idx, spans["규격"][0], matched.request.규격)
        if "단위" in spans:
            sheet.write(row_idx, spans["단위"][0], matched.request.단위)
        if "수량" in spans:
            sheet.write(row_idx, spans["수량"][0], matched.request.구매량 or "")
        if "단가" in spans:
            sheet.write(row_idx, spans["단가"][0], matched.단가 or "")
        if "금액" in spans:
            sheet.write(row_idx, spans["금액"][0], matched.금액 or "")


def generate_estimate(
    db_path: Path,
    input_path: Path,
    output_path: Path,
    template_path: Optional[Path],
    overrides: Optional[Dict[int, Optional[float]]] = None,
    request_label: Optional[str] = None,
) -> Path:
    if template_path is None:
        if input_path.is_dir():
            template_path = pick_latest_template(input_path)
        else:
            template_path = pick_latest_template(db_path.parent)

    matched = match_prices(db_path, parse_request(input_path))
    if overrides:
        for idx, item in enumerate(matched):
            if idx not in overrides:
                continue
            unit_price = overrides[idx]
            if unit_price is None:
                item.단가 = None
                item.금액 = None
            else:
                item.단가 = float(unit_price)
                item.금액 = (
                    item.단가 * item.request.구매량
                    if item.request.구매량 is not None
                    else None
                )
    if not request_label:
        request_label = compute_request_label(input_path)

    if template_path.suffix.lower() == ".xlsx":
        try:
            wb = load_workbook(template_path)
        except BadZipFile as exc:
            raise SystemExit(
                "템플릿 파일이 올바른 .xlsx 형식이 아닙니다. 엑셀에서 다시 저장해 주세요."
            ) from exc
        if not wb.worksheets:
            raise SystemExit(
                "템플릿에 시트가 없습니다. 엑셀에서 다시 저장해 주세요."
            )
        target_sheet = None
        header_row = None
        header_map: Dict[str, int] = {}
        candidates = []
        for ws in wb.worksheets:
            header_info = find_header_row_openpyxl(ws)
            if not header_info:
                continue
            row_idx, map_info = header_info
            non_empty = 0
            for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=25):
                for cell in row:
                    if cell.value not in (None, ""):
                        non_empty += 1
            score = len(ws.merged_cells.ranges) + non_empty
            candidates.append((score, ws, row_idx, map_info))
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            _, target_sheet, header_row, header_map = candidates[0]
        if target_sheet is None or header_row is None:
            target_sheet = wb.worksheets[0]
            header_row = 13
            header_map = {
                "번호": 1,
                "품명": 2,
                "규격": 8,
                "단위": 12,
                "수량": 15,
                "단가": 16,
                "금액": 20,
            }
        wb.active = wb.worksheets.index(target_sheet)

        if request_label:
            target_sheet.cell(row=7, column=1).value = request_label

        start_row = header_row + 1
        start_col = 2
        end_col = 24
        last_data_row = find_last_data_row(target_sheet, start_row, start_col, end_col)
        final_row = max(last_data_row, start_row + len(matched) - 1)
        for row_idx in range(start_row, final_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = target_sheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        for idx, item in enumerate(matched, start=1):
            row_idx = start_row + (idx - 1)
            if "번호" in header_map:
                resolve_writable_cell(
                    target_sheet, row_idx, header_map["번호"]
                ).value = idx
            resolve_writable_cell(
                target_sheet, row_idx, header_map["품명"]
            ).value = item.request.품명
            resolve_writable_cell(
                target_sheet, row_idx, header_map["규격"]
            ).value = item.request.규격
            resolve_writable_cell(
                target_sheet, row_idx, header_map["단위"]
            ).value = item.request.단위
            resolve_writable_cell(
                target_sheet, row_idx, header_map["수량"]
            ).value = item.request.구매량 if item.request.구매량 is not None else ""
            resolve_writable_cell(
                target_sheet, row_idx, header_map["단가"]
            ).value = item.단가 if item.단가 is not None else ""
            resolve_writable_cell(
                target_sheet, row_idx, header_map["금액"]
            ).value = item.금액 if item.금액 is not None else ""

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    book = xlrd.open_workbook(template_path)
    writable = xl_copy(book)
    sheet = writable.get_sheet(0)
    sheet.write(6, 0, request_label)

    df = pd.read_excel(template_path, sheet_name=0, header=None)
    header_info = find_header_row(df)
    if not header_info:
        raise SystemExit("템플릿에서 헤더 행을 찾지 못했습니다.")
    header_row, _ = header_info
    spans = compute_spans(df, header_row)
    last_row = len(df) - 1
    clear_existing_rows(sheet, header_row, spans, last_row)
    write_rows(sheet, header_row + 1, spans, matched)

    if output_path.suffix.lower() != ".xls":
        output_path = output_path.with_suffix(".xls")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    writable.save(str(output_path))
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Level 4: 신규 견적서 생성")
    parser.add_argument("db_path", help="Level 1에서 생성한 SQLite DB 경로")
    parser.add_argument("input_path", help="견적의뢰서 파일 또는 DB 폴더 경로")
    parser.add_argument("output_path", help="생성할 견적서 파일 경로 (.xls)")
    parser.add_argument(
        "--template",
        help="기준 견적서 템플릿 경로 (미지정 시 최신 견적서 사용)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    db_path = Path(args.db_path).expanduser().resolve()
    input_path = Path(args.input_path).expanduser().resolve()
    output_path = Path(args.output_path).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve() if args.template else None

    result = generate_estimate(db_path, input_path, output_path, template_path)
    print(f"완료: {result}")


if __name__ == "__main__":
    main()
